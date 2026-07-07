"""
Spec Excel -> HTML transposer.

Reads a spec Excel file (e.g. 200.xlsx) populated against the standard yard
spec template and emits a corresponding HTML file (spec_<jobno>.html) built on
spec-template.html.

Design notes:
  * Pure string / regex edits on the HTML — no DOM reformatting. Functional
    JS attributes (ids, data-* attributes, classes) are preserved byte-for-byte
    everywhere we don't intentionally touch.
  * Yard Quote rows are pre-rendered. The pagination engine treats `.yd-section`
    as one opaque block measured by rendered height — adding rows just makes
    the section taller, which the engine handles via its existing overflow
    loop. No JS row-count assumption to disturb (verified — see chat).
  * The 5 default `<tr>` rows in the template are reused in place for items
    1–5; rows 6+ are appended as clones with sequential `data-yd-*` indices.
    For specs with fewer than 5 items, surplus template rows are left empty.
  * Field-set rules per spec (per user direction):
      - Vessel Name, IMO, Class/RO, Department: not transposed (external static)
      - Equipment Type, Equipment Serial: leave blank
      - Spares/Services scope, Job Status, Completion Date: no changes required
      - HTML-only Yes/No toggles and group-header behaviour: untouched
  * Inclusion checkboxes: state mirrored from Excel form-control state. Items
    in HTML with no Excel control (Lifting equipment, Underwriters, Maker's
    supply, Sample) default to unchecked.
  * Full Description is the merged content of every populated cell in cols
    B–H between the "Specific Job Details" header row and the
    "YARD TO QUOTE..." row, top-to-bottom.
"""

from __future__ import annotations
import re
import sys
import zipfile
from pathlib import Path
from html import escape as html_escape
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore")  # quiet openpyxl drawing warnings


# ----------------------------------------------------------------------------
# EXCEL READING
# ----------------------------------------------------------------------------

# Inclusion labels that are section headers (skip when matching checkboxes)
SECTION_HEADERS = {
    "YARD WORK", "SUB-CONTRACT", "OTHER DETAILS",
    "INSPECTIONS / SURVEYS", "MATERIAL", "ENCLOSURES",
}


def _normalize_label(s: str) -> str:
    """Normalize an inclusion label for comparison: trim, lower-case, unify
    apostrophe variants (curly/straight), and ampersand spacing."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("\u2019", "'").replace("&amp;", "&")
    s = re.sub(r"\s+", " ", s)
    return s


def find_row_by_marker(ws, col: int, marker: str) -> int | None:
    """Return 1-indexed row number of first cell in `col` whose value contains `marker`."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v and marker.lower() in str(v).lower():
            return r
    return None


# ----------------------------------------------------------------------------
# SJD STRUCTURE DETECTION
# Spec 406 introduced tabulated content inside "Specific Job Details" — a
# 3-column block (label / qty / optional note) embedded between paragraph rows.
# We classify each SJD row by its merge pattern:
#   * B is master of B:H (or wider)  → free-text paragraph row
#   * B is master of B:D + E:G has content + F empty  → table HEADER row
#   * B is master of B:D (with or without F qty / G:H note)  → table DATA row
#   * Anything else  → treat as paragraph
# A spec is "SJD-tabulated" iff at least one table row is detected.
# ----------------------------------------------------------------------------

class SjdTabulationDetected(Exception):
    """Raised when an SJD table is detected in batch (auto) mode."""

    def __init__(self, xlsx_path, table_row_count):
        self.xlsx_path = xlsx_path
        self.table_row_count = table_row_count
        super().__init__(
            f"SJD tabulation detected in {xlsx_path} ({table_row_count} table rows). "
            f"Skipping per batch policy — re-run with --allow-sjd-table to process."
        )


class NonStandardYardQuoteDetected(Exception):
    """Raised when the Yard-Quote table uses a non-standard structure (multi-
    character SL codes, e.g. spec 405's `A1, A2, ..., B1, B2` layout grouped
    under section headers)."""

    def __init__(self, xlsx_path, sample_sl, count):
        self.xlsx_path = xlsx_path
        self.sample_sl = sample_sl
        self.count = count
        super().__init__(
            f"Non-standard Yard Quote structure detected in {xlsx_path} "
            f"({count} multi-character SL codes such as {sample_sl!r}). "
            f"Skipping per batch policy — re-run with --allow-nonstandard-yard "
            f"to process (case-by-case)."
        )


# A "standard" SL code in Yard Quote is a single A–Z letter. Multi-character
# alphanumeric codes like A1, B2, AA1 indicate a grouped/sectioned variant
# (e.g. spec 405) that needs case-by-case handling.
NONSTANDARD_SL_PATTERN = re.compile(r"^[A-Z]+\d+$")


def detect_yard_quote_variant(ws) -> dict:
    """Inspect the Yard Quote region. Returns:
      {'kind': 'standard' | 'nonstandard', 'sample_sl': str|None, 'count': int}
    where 'count' is the number of non-standard SL codes seen."""
    header_row = find_row_by_marker(ws, 2, "YARD TO QUOTE")
    if header_row is None:
        return {"kind": "standard", "sample_sl": None, "count": 0}

    nonstd: list[str] = []
    for r in range(header_row + 2, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        b_s = str(b).strip() if b is not None else ""
        if b_s.lower() == "scope":
            break
        if not b_s:
            continue
        if NONSTANDARD_SL_PATTERN.match(b_s):
            nonstd.append(b_s)

    if nonstd:
        return {"kind": "nonstandard", "sample_sl": nonstd[0], "count": len(nonstd)}
    return {"kind": "standard", "sample_sl": None, "count": 0}


def _build_merge_extents(ws, row_min: int, row_max: int) -> dict:
    """Map (row, col) -> (master_row, master_col, max_col) for every cell in
    a merged range that intersects [row_min, row_max]."""
    out = {}
    for mr in ws.merged_cells.ranges:
        if mr.max_row < row_min or mr.min_row > row_max:
            continue
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                out[(r, c)] = (mr.min_row, mr.min_col, mr.max_col)
    return out


def _classify_sjd_row(ws, r: int, merges: dict) -> dict:
    """Classify a single SJD row. See module-level comment for the rules."""
    b_val = ws.cell(row=r, column=2).value
    b_str = str(b_val).strip() if b_val is not None else ""
    if not b_str:
        return {"kind": "empty"}

    b_master = merges.get((r, 2))
    b_max_col = b_master[2] if b_master else 2

    # Wide merge (B extends to G or beyond) → paragraph
    if b_master and b_max_col >= 7:
        return {"kind": "paragraph", "text": b_str}

    # Narrow merge (B:D, ending at col 4)
    if b_master and b_max_col == 4:
        # Header row signature: E:G merge with content + empty F
        e_val = ws.cell(row=r, column=5).value
        e_master = merges.get((r, 5))
        f_val = ws.cell(row=r, column=6).value
        if e_master and e_val and not f_val:
            return {
                "kind": "table_header",
                "title": b_str,
                "qty_header": str(e_val).strip(),
            }
        # Otherwise table data row
        f_str = str(f_val).strip() if f_val is not None else ""
        g_val = ws.cell(row=r, column=7).value
        g_str = str(g_val).strip() if g_val is not None else ""
        return {"kind": "table_row", "label": b_str, "qty": f_str, "note": g_str}

    # B has unexpected merge / no merge → fall through as paragraph best-effort
    return {"kind": "paragraph", "text": b_str}


def extract_sjd_layout(ws) -> dict:
    """Walk the SJD region and return a structured layout:
      {
        'has_table':         bool,
        'table_title':       str | None,   # e.g. "Following are details of Ship's Marks:"
        'table_qty_header':  str | None,   # e.g. "Number of Locations"
        'table_rows':        [{'label', 'qty', 'note'}, ...],
        'free_text_before':  [str, ...],   # paragraphs before the table
        'free_text_after':   [str, ...],   # paragraphs after the table
        'free_text_all':     str,          # convenience: all paragraphs joined for textarea
      }
    For non-tabulated specs (e.g. 200), has_table=False and free_text_all is
    the same as the legacy extract_full_description() output.
    """
    sjd_start = find_row_by_marker(ws, 2, "Specific Job Details")
    yto = find_row_by_marker(ws, 2, "YARD TO QUOTE")
    if sjd_start is None or yto is None:
        return {
            "has_table": False, "table_title": None, "table_qty_header": None,
            "table_rows": [], "free_text_before": [], "free_text_after": [],
            "free_text_all": "",
        }

    merges = _build_merge_extents(ws, sjd_start, yto)

    table_title = None
    table_qty_header = None
    table_rows: list[dict] = []
    free_before: list[str] = []
    free_after: list[str] = []
    saw_table = False

    for r in range(sjd_start + 1, yto):
        info = _classify_sjd_row(ws, r, merges)
        kind = info["kind"]
        if kind == "empty":
            continue
        if kind == "paragraph":
            (free_after if saw_table else free_before).append(info["text"])
        elif kind == "table_header":
            saw_table = True
            table_title = info["title"]
            table_qty_header = info["qty_header"]
        elif kind == "table_row":
            saw_table = True
            table_rows.append({
                "label": info["label"],
                "qty":   info["qty"],
                "note":  info["note"],
            })

    has_table = bool(table_rows or table_title)
    return {
        "has_table": has_table,
        "table_title": table_title,
        "table_qty_header": table_qty_header,
        "table_rows": table_rows,
        "free_text_before": free_before,
        "free_text_after": free_after,
        "free_text_all": "\n\n".join(free_before + free_after),
    }


def extract_full_description(ws) -> str:
    """Concatenate every populated cell in cols B–H between the
    'Specific Job Details' header row and the 'YARD TO QUOTE' row."""
    start_marker_row = find_row_by_marker(ws, 2, "Specific Job Details")
    end_marker_row = find_row_by_marker(ws, 2, "YARD TO QUOTE")
    if start_marker_row is None or end_marker_row is None:
        return ""

    blocks = []
    for r in range(start_marker_row + 1, end_marker_row):
        row_pieces = []
        for c in range(2, 9):  # B–H
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                row_pieces.append(s)
        if row_pieces:
            blocks.append(" ".join(row_pieces))
    return "\n\n".join(blocks)


def extract_yard_quote(ws) -> list[tuple[str, str]]:
    """Return [(SL_letter, description), ...] from the 'YARD TO QUOTE' table.
    Continuation rows (B empty, C populated) are merged into the prior item."""
    header_row = find_row_by_marker(ws, 2, "YARD TO QUOTE")
    if header_row is None:
        return []

    # Skip the column-label row (SL./JOB DECRIPTION/YARD QUOTE) at header_row+1.
    # Items begin at header_row+2 and run until we hit "Scope" in column B.
    items: list[tuple[str, list[str]]] = []
    for r in range(header_row + 2, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        b_s = str(b).strip() if b is not None else ""
        c_s = str(c).strip() if c is not None else ""

        # Stop at the "Scope" table that follows
        if b_s.lower() == "scope":
            break

        if not b_s and not c_s:
            continue  # blank spacer row

        if len(b_s) == 1 and b_s.isalpha():
            items.append((b_s, [c_s] if c_s else []))
        else:
            # Continuation row — append to last item's description
            if items and c_s:
                items[-1][1].append(c_s)

    return [(sl, " ".join(parts).strip()) for sl, parts in items]


def extract_form_control_states(xlsx_path: Path) -> list[dict]:
    """Read embedded Excel form-control checkboxes. Returns list of
    {row, col, checked} (1-indexed cell anchor) for each control."""
    with zipfile.ZipFile(xlsx_path) as z:
        # Map relationship id -> ctrlProp filename
        try:
            rels = z.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8", "ignore")
        except KeyError:
            return []
        rid_to_ctrl = {
            m.group(1): m.group(2)
            for m in re.finditer(
                r'Id="([^"]+)"\s+Type="[^"]*ctrlProp"\s+Target="\.\./ctrlProps/([^"]+)"',
                rels,
            )
        }

        # Read each ctrlProp for checked state
        ctrl_states: dict[str, bool] = {}
        for name in z.namelist():
            if name.startswith("xl/ctrlProps/") and name.endswith(".xml"):
                xml = z.read(name).decode("utf-8", "ignore")
                # checkbox state: <formControlPr ... checked="Checked"/>
                ctrl_states[name.rsplit("/", 1)[-1]] = "checked" in xml.lower()

        sheet = z.read("xl/worksheets/sheet1.xml").decode("utf-8", "ignore")

    out = []
    for m in re.finditer(
        r'<control\s+shapeId="\d+"\s+r:id="([^"]+)"\s+name="[^"]+"[^>]*>(.*?)</control>',
        sheet, re.DOTALL,
    ):
        rid, body = m.group(1), m.group(2)
        anchor = re.search(
            r"<from>\s*<xdr:col>(\d+)</xdr:col>\s*<xdr:colOff>[^<]*</xdr:colOff>\s*<xdr:row>(\d+)</xdr:row>",
            body,
        )
        if not anchor:
            continue
        ctrl_file = rid_to_ctrl.get(rid)
        if ctrl_file is None:
            continue
        out.append({
            "row": int(anchor.group(2)) + 1,  # XML is 0-indexed; convert
            "col": int(anchor.group(1)) + 1,
            "checked": ctrl_states.get(ctrl_file, False),
        })
    return out


def extract_checked_inclusions(ws, controls: list[dict]) -> set[str]:
    """For each form control, look up the inclusion label in column K at the
    same row. Return the set of normalized labels whose box is checked."""
    checked = set()
    for ctrl in controls:
        if not ctrl["checked"]:
            continue
        label_cell = ws.cell(row=ctrl["row"], column=11)  # column K
        if label_cell.value is None:
            continue
        label = str(label_cell.value).strip()
        if label.upper() in SECTION_HEADERS:
            continue
        checked.add(_normalize_label(label))
    return checked


def extract_spec_data(xlsx_path: Path) -> dict:
    """One-stop reader: returns a dict of all transposable values from a spec."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    def cell(addr):
        v = ws[addr].value
        return None if v is None else str(v).strip()

    return {
        "job_id":         cell("J5"),                      # Office job number
        "done_by":        cell("D67"),                     # SHIP-YARD / OWNER's CREW / etc.
        "done_when":      cell("E67"),                     # During Docking / Anytime
        "job_desc_brief": cell("C8"),                      # DESCR. value (e.g. "DOCKING GENERAL")
        "equip_maker":    cell("C12"),
        "equip_model":    cell("C13"),
        "full_desc":      extract_full_description(ws),
        "yard_quote":     extract_yard_quote(ws),
        "spares_status":  cell("D64"),
        "spares_po":      cell("E64"),
        "services_status": cell("D65"),
        "services_po":    cell("E65"),
        "job_status":     cell("D66"),
        "checked_inclusions": extract_checked_inclusions(
            ws, extract_form_control_states(xlsx_path)
        ),
    }


# ----------------------------------------------------------------------------
# HTML PATCHING
# ----------------------------------------------------------------------------

def set_input_value(html: str, input_id: str, value: str | None) -> str:
    """Add or replace value="..." on <input id="...">. No-op if value is None."""
    if value is None:
        return html
    pattern = re.compile(
        r'(<input\b[^>]*\bid="' + re.escape(input_id) + r'"[^>]*?)(\s*/?>)'
    )

    def repl(m: re.Match) -> str:
        before, end = m.group(1), m.group(2)
        before = re.sub(r'\s+value="[^"]*"', "", before)  # drop existing
        return before + ' value="' + html_escape(value, quote=True) + '"' + end

    new, n = pattern.subn(repl, html, count=1)
    if n == 0:
        raise RuntimeError(f"input id={input_id!r} not found")
    return new


def set_textarea_content(html: str, textarea_id: str, content: str | None) -> str:
    """Replace the body of <textarea id="..."></textarea>."""
    if content is None:
        return html
    pattern = re.compile(
        r'(<textarea\b[^>]*\bid="' + re.escape(textarea_id) + r'"[^>]*>)(.*?)(</textarea>)',
        re.DOTALL,
    )
    new, n = pattern.subn(lambda m: m.group(1) + html_escape(content) + m.group(3), html, count=1)
    if n == 0:
        raise RuntimeError(f"textarea id={textarea_id!r} not found")
    return new


def set_select_by_text(html: str, select_id: str, target_text: str | None) -> str:
    """Add `selected` to the <option> inside <select id="..."> whose visible
    text matches target_text (case-insensitive, trimmed). No-op if target_text
    is None or empty. Idempotent — strips any prior `selected` first."""
    if not target_text:
        return html
    target_norm = target_text.strip().lower()

    select_pat = re.compile(
        r'(<select\b[^>]*\bid="' + re.escape(select_id) + r'"[^>]*>)(.*?)(</select>)',
        re.DOTALL,
    )

    def repl_select(sm: re.Match) -> str:
        opening, body, closing = sm.group(1), sm.group(2), sm.group(3)
        # Drop any existing `selected` markers in this select
        body = re.sub(r'(\s+)selected(?:="[^"]*")?', "", body)

        # Match each <option ...>TEXT</option> and inject `selected` on a hit
        def repl_opt(om: re.Match) -> str:
            attrs = om.group(1)
            text = om.group(2)
            if text.strip().lower() == target_norm:
                return f"<option{attrs} selected>{text}</option>"
            return om.group(0)

        body = re.sub(r"<option(\s+[^>]*)?>([^<]*)</option>", repl_opt, body, flags=re.IGNORECASE)
        return opening + body + closing

    new, n = select_pat.subn(repl_select, html, count=1)
    if n == 0:
        raise RuntimeError(f"select id={select_id!r} not found")
    return new


def set_select_by_attr(html: str, select_attr: str, attr_value: str,
                       target_text: str | None) -> str:
    """Like set_select_by_text but matches the <select> by `[attr="value"]`
    (used for the Spares/Services scope rows which use `data-sp-status`)."""
    if not target_text:
        return html
    target_norm = target_text.strip().lower()

    select_pat = re.compile(
        r'(<select\b[^>]*\b' + re.escape(select_attr) + r'="' + re.escape(attr_value) + r'"[^>]*>)(.*?)(</select>)',
        re.DOTALL,
    )

    def repl_select(sm: re.Match) -> str:
        opening, body, closing = sm.group(1), sm.group(2), sm.group(3)
        body = re.sub(r'(\s+)selected(?:="[^"]*")?', "", body)

        def repl_opt(om: re.Match) -> str:
            attrs = om.group(1) or ""
            text = om.group(2)
            if text.strip().lower() == target_norm:
                return f"<option{attrs} selected>{text}</option>"
            return om.group(0)

        body = re.sub(r"<option(\s+[^>]*)?>([^<]*)</option>", repl_opt, body, flags=re.IGNORECASE)
        return opening + body + closing

    new, n = select_pat.subn(repl_select, html, count=1)
    if n == 0:
        raise RuntimeError(f"select [{select_attr}={attr_value!r}] not found")
    return new


def set_input_by_attr(html: str, attr: str, attr_value: str, value: str | None) -> str:
    """Set value="..." on <input ... [attr="attr_value"] ...>. Used for
    `data-sp-po="spares"` and `data-sp-po="services"`."""
    if value is None:
        return html
    pattern = re.compile(
        r'(<input\b[^>]*\b' + re.escape(attr) + r'="' + re.escape(attr_value) + r'"[^>]*?)(\s*/?>)'
    )

    def repl(m: re.Match) -> str:
        before, end = m.group(1), m.group(2)
        before = re.sub(r'\s+value="[^"]*"', "", before)
        return before + ' value="' + html_escape(value, quote=True) + '"' + end

    new, n = pattern.subn(repl, html, count=1)
    if n == 0:
        raise RuntimeError(f"input [{attr}={attr_value!r}] not found")
    return new


def normalize_status_value(value: str | None) -> str | None:
    """Normalize free-form status text to template option labels."""
    if not value:
        return value
    key = value.strip().lower()
    canon = {
        "pending": "Pending",
        "in progress": "In Progress",
        "completed": "Completed",
        "n/a": "N/A",
        "na": "N/A",
    }
    return canon.get(key, value)


def render_yard_quote(html: str, items: list[tuple[str, str]]) -> str:
    """Replace the contents of the .yd-table <tbody> with one <tr> per item.
    For < 5 items, pad with empty rows up to 5 (matches default template).
    For > 5 items, append additional rows with sequential data-yd-* indices."""
    # Locate the yd-table tbody
    tbody_pat = re.compile(
        r'(<table\s+class="yd-table"[^>]*>.*?<tbody>)(.*?)(</tbody>)',
        re.DOTALL,
    )
    m = tbody_pat.search(html)
    if not m:
        raise RuntimeError("yd-table tbody not found")

    n_rows = max(len(items), 5)  # keep at least 5 rows so default template feel is preserved
    rows_html = []
    for i in range(1, n_rows + 1):
        if i <= len(items):
            sl_letter, desc = items[i - 1]
        else:
            sl_letter, desc = ("", "")
        sl_attr = html_escape(sl_letter, quote=True)
        desc_html = html_escape(desc)
        rows_html.append(
            f'                <tr>\n'
            f'                  <td><input class="yd-sl-input" type="text" value="{sl_attr}" data-yd-sl="{i}" /></td>\n'
            f'                  <td><textarea class="yd-job-textarea" rows="1" data-yd-row="{i}">{desc_html}</textarea></td>\n'
            f'                  <td><input class="yd-quote-input" type="text" data-yd-quote="{i}" maxlength="15" /></td>\n'
            f'                </tr>'
        )

    new_tbody = "\n" + "\n".join(rows_html) + "\n              "
    return tbody_pat.sub(lambda mm: mm.group(1) + new_tbody + mm.group(3), html, count=1)


def fix_print_button(html: str) -> str:
    """Preserve template print wiring unchanged.

    The template already has a delegated click handler for `#btnPrintAction`.
    Keeping generated output byte-close to the template avoids print-flow
    regressions tied to inline onclick differences.
    """
    return html


def check_inclusion_boxes(html: str, checked_labels: set[str]) -> str:
    """For each <label class="inclusion-row">...LABEL_TEXT</label>, if the
    normalized LABEL_TEXT is in checked_labels, add `checked` to the inner
    <input type="checkbox" />."""
    pattern = re.compile(
        r'(<label\s+class="inclusion-row">\s*<input\s+type="checkbox"\s*)(/>)(\s*)([^<]+)(</label>)'
    )

    def repl(m: re.Match) -> str:
        pre, slash, gap, text, end = m.groups()
        if _normalize_label(text) in checked_labels:
            return f'{pre}checked {slash}{gap}{text}{end}'
        return m.group(0)

    return pattern.sub(repl, html)


# ----------------------------------------------------------------------------
# SJD TABLE SECTION — HTML / CSS GENERATION & INJECTION
# ----------------------------------------------------------------------------

SJD_TABLE_CSS = """\
    /* SJD Table Section — injected by transpose-spec.py for specs with
       tabulated content inside Specific Job Details. Sits as a regular
       .flow-box between Equipment Details and Job Details, so the existing
       pagination engine treats it as an opaque block. */
    .sjd-table-section {
      width: calc(100% - 2px);
      max-width: calc(100% - 2px);
      border: 1px solid var(--section-box-border, #cfd2d6);
      border-radius: 2mm;
      background: var(--section-box-bg, #ffffff);
      padding: 6px 8px;
      box-sizing: border-box;
      margin-top: 5px;
    }
    .sjd-table-header { margin-bottom: 6px; }
    .sjd-table-title {
      display: block;
      font-weight: 600;
      font-size: 12px;
      color: #111827;
    }
    .sjd-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 11px;
      table-layout: fixed;
    }
    .sjd-col-label { width: 50%; }
    .sjd-col-qty   { width: 15%; }
    .sjd-col-note  { width: 35%; }
    .sjd-table thead th {
      background: #e8eef5;
      border: 1px solid #c9d1da;
      padding: 4px 6px;
      font-weight: 600;
      font-size: 10px;
      text-align: center;
      line-height: 1.2;
    }
    .sjd-table tbody td {
      border: 1px solid #d9d9d9;
      padding: 0;
      vertical-align: middle;
    }
    .sjd-label-input,
    .sjd-qty-input {
      width: 100%;
      box-sizing: border-box;
      border: 0;
      padding: 3px 6px;
      font: inherit;
      font-size: 11px;
      background: transparent;
      color: #111827;
      line-height: 1.4;
    }
    .sjd-qty-input { text-align: center; }
    .sjd-label-input::placeholder,
    .sjd-qty-input::placeholder {
      color: #9ca3af;
      font-style: italic;
    }
    .sjd-label-input:focus,
    .sjd-qty-input:focus {
      background: #fff8e1;
      outline: 1px solid #facc15;
      outline-offset: -1px;
    }
    .sjd-note {
      display: block;
      padding: 3px 6px;
      color: #6b7280;
      font-size: 10px;
      font-style: italic;
      line-height: 1.3;
    }
"""

# Patterns used to recognise placeholder rows imported from Excel
PLACEHOLDER_TOKENS = {"[placeholder]", "[Placeholder]", "[PLACEHOLDER]", "<placeholder>"}


def _is_placeholder_row(row: dict) -> bool:
    """Whether this row's label should render as an empty input (Reading B)
    rather than a pre-filled input."""
    label = (row.get("label") or "").strip()
    return (
        label in PLACEHOLDER_TOKENS
        or label.lower() == "[placeholder]"
        or (label.startswith("[") and label.endswith("]") and "placeholder" in label.lower())
    )


def generate_sjd_table_section(layout: dict) -> str:
    """Render the .flow-box.sjd-table-section HTML for one spec.
    `layout` is the dict returned by extract_sjd_layout()."""
    rows_html = []
    for i, row in enumerate(layout["table_rows"], 1):
        label = row.get("label", "") or ""
        qty   = row.get("qty", "") or ""
        note  = row.get("note", "") or ""

        if _is_placeholder_row(row):
            label_attr = ' value="" placeholder="(user-defined item)"'
            qty_attr   = ' value="" placeholder="qty"'
        else:
            label_attr = f' value="{html_escape(label, quote=True)}"'
            qty_attr   = f' value="{html_escape(qty,   quote=True)}"'

        rows_html.append(
            f'              <tr>\n'
            f'                <td><input class="sjd-label-input" type="text"{label_attr} data-sjd-row="{i}" /></td>\n'
            f'                <td><input class="sjd-qty-input"   type="text"{qty_attr}   data-sjd-row="{i}" /></td>\n'
            f'                <td><span class="sjd-note">{html_escape(note)}</span></td>\n'
            f'              </tr>'
        )

    title   = html_escape(layout.get("table_title")      or "")
    qty_hdr = html_escape(layout.get("table_qty_header") or "")

    body = "\n".join(rows_html)
    return (
        '        <section class="flow-box sjd-table-section">\n'
        '          <div class="sjd-table-header">\n'
        f'            <span class="sjd-table-title">{title}</span>\n'
        '          </div>\n'
        '          <table class="sjd-table">\n'
        '            <colgroup>\n'
        '              <col class="sjd-col-label" />\n'
        '              <col class="sjd-col-qty" />\n'
        '              <col class="sjd-col-note" />\n'
        '            </colgroup>\n'
        '            <thead>\n'
        '              <tr>\n'
        '                <th class="sjd-th-label"></th>\n'
        f'                <th class="sjd-th-qty">{qty_hdr}</th>\n'
        '                <th class="sjd-th-note"></th>\n'
        '              </tr>\n'
        '            </thead>\n'
        '            <tbody>\n'
        f'{body}\n'
        '            </tbody>\n'
        '          </table>\n'
        '        </section>\n'
    )


def inject_sjd_table_styles(html: str) -> str:
    """Append a <style id="sjd-table-styles"> block before </head>.
    Idempotent — won't double-insert."""
    if 'id="sjd-table-styles"' in html:
        return html
    block = f'  <style id="sjd-table-styles">\n{SJD_TABLE_CSS}  </style>\n'
    new, n = re.subn(r"</head>", block + "</head>", html, count=1)
    if n == 0:
        raise RuntimeError("</head> not found — cannot inject SJD styles")
    return new


def inject_sjd_table_section(html: str, section_html: str) -> str:
    """Insert the .sjd-table-section flow-box immediately before the
    .job-details-section flow-box on page 2."""
    pattern = re.compile(
        r'([ \t]*)(<section\s+class="flow-box\s+job-details-section)'
    )
    new, n = pattern.subn(
        lambda m: section_html + m.group(1) + m.group(2),
        html,
        count=1,
    )
    if n == 0:
        raise RuntimeError("job-details-section anchor not found — cannot inject SJD table")
    return new


# ----------------------------------------------------------------------------
# MAIN ENTRY
# ----------------------------------------------------------------------------
def transpose(xlsx_path: Path, template_path: Path, output_path: Path,
              allow_sjd_table: bool = False,
              allow_nonstandard_yard: bool = False) -> dict:
    """Run the full transposition. Returns the data dict for inspection.

    If the spec's Specific Job Details contains a tabulation (multi-column
    structured data, e.g. spec 406's Ship's Marks layout) and `allow_sjd_table`
    is False, raises SjdTabulationDetected so a batch caller can skip & log.
    With allow_sjd_table=True, generates an interactive table flow-box section
    inserted before the Job Details textarea (Option 3 + Reading B).

    If the spec's Yard Quote uses non-standard SL codes (multi-character like
    A1, B2 — e.g. spec 405's grouped-by-section layout) and
    `allow_nonstandard_yard` is False, raises NonStandardYardQuoteDetected."""
    data = extract_spec_data(xlsx_path)

    # Inspect SJD layout up-front so we can decide whether to skip
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sjd_layout = extract_sjd_layout(ws)

    if sjd_layout["has_table"] and not allow_sjd_table:
        raise SjdTabulationDetected(xlsx_path, len(sjd_layout["table_rows"]))

    # Inspect Yard Quote variant
    yq_variant = detect_yard_quote_variant(ws)
    if yq_variant["kind"] == "nonstandard" and not allow_nonstandard_yard:
        raise NonStandardYardQuoteDetected(
            xlsx_path, yq_variant["sample_sl"], yq_variant["count"]
        )

    # If we're processing a tabulated spec, the textarea content is just the
    # paragraph rows (the table content goes into its own flow-box). Otherwise
    # use the legacy full-description merge.
    full_desc = sjd_layout["free_text_all"] if sjd_layout["has_table"] else data["full_desc"]

    html = template_path.read_text(encoding="utf-8")

    # Basic input fields
    html = set_input_value(html, "specJobId",   data["job_id"])
    html = set_input_value(html, "specDoneBy",  data["done_by"])
    html = set_input_value(html, "specDoneWhen", data["done_when"])
    html = set_input_value(html, "specJobDesc", data["job_desc_brief"])
    html = set_input_value(html, "equipMaker",  data["equip_maker"])
    html = set_input_value(html, "equipModel",  data["equip_model"])

    # Full description (textarea) — paragraphs only when an SJD table exists
    html = set_textarea_content(html, "jobFullDesc", full_desc)

    # Yard Quote table — pre-rendered rows
    html = render_yard_quote(html, data["yard_quote"])

    # Scope table (Spares / Services) — selects + PO inputs
    html = set_select_by_attr(html, "data-sp-status", "spares",   normalize_status_value(data["spares_status"]))
    html = set_input_by_attr (html, "data-sp-po",     "spares",   data["spares_po"])
    html = set_select_by_attr(html, "data-sp-status", "services", normalize_status_value(data["services_status"]))
    html = set_input_by_attr (html, "data-sp-po",     "services", data["services_po"])

    # Job Status select — capitalize first letter (Excel uses "PENDING")
    job_status = normalize_status_value(data["job_status"])
    html = set_select_by_text(html, "jobStatusField", job_status)

    # Inclusion checkboxes
    html = check_inclusion_boxes(html, data["checked_inclusions"])

    # SJD table section (only when present + allowed)
    if sjd_layout["has_table"]:
        html = inject_sjd_table_styles(html)
        html = inject_sjd_table_section(html, generate_sjd_table_section(sjd_layout))

    # Keep template-native print behavior in generated output.
    html = fix_print_button(html)

    output_path.write_text(html, encoding="utf-8")

    # Return both data and layout for caller inspection / reporting
    return {**data, "_sjd_layout": sjd_layout}


def main(argv: list[str]) -> int:
    args = list(argv[1:])
    allow_sjd_table = False
    allow_nonstandard_yard = False
    if "--allow-sjd-table" in args:
        args.remove("--allow-sjd-table")
        allow_sjd_table = True
    if "--allow-nonstandard-yard" in args:
        args.remove("--allow-nonstandard-yard")
        allow_nonstandard_yard = True

    if not args:
        print("Usage: transpose-spec.py [--allow-sjd-table] [--allow-nonstandard-yard] "
              "<spec.xlsx> [template.html] [output.html]")
        return 1

    xlsx_path = Path(args[0])
    template_path = Path(args[1]) if len(args) > 1 else Path("spec-template.html")

    if len(args) > 2:
        output_path = Path(args[2])
    else:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        job_no = ws["J5"].value
        wb.close()
        job_token = str(job_no).strip() if job_no is not None else "unknown"
        output_path = Path("outputs") / f"Spec-{job_token}.html"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        data = transpose(xlsx_path, template_path, output_path,
                         allow_sjd_table=allow_sjd_table,
                         allow_nonstandard_yard=allow_nonstandard_yard)
    except SjdTabulationDetected as e:
        print(f"⚠️  SKIPPED: {e}")
        return 2
    except NonStandardYardQuoteDetected as e:
        print(f"⚠️  SKIPPED: {e}")
        return 3

    print(f"Wrote {output_path}")
    print(f"  Job ID:           {data['job_id']!r}")
    print(f"  Done by / when:   {data['done_by']!r} / {data['done_when']!r}")
    print(f"  Brief desc:       {data['job_desc_brief']!r}")
    print(f"  Maker / Model:    {data['equip_maker']!r} / {data['equip_model']!r}")

    sjd = data.get("_sjd_layout") or {}
    if sjd.get("has_table"):
        print(f"  SJD table:        {len(sjd['table_rows'])} rows "
              f"(title={sjd['table_title']!r}, qty_hdr={sjd['table_qty_header']!r})")
        print(f"  Free-text rows:   {len(sjd['free_text_before'])} before + {len(sjd['free_text_after'])} after table")
    else:
        print(f"  Full desc length: {len(data['full_desc'])} chars")

    print(f"  Yard quote items: {len(data['yard_quote'])}")
    for sl, desc in data["yard_quote"]:
        print(f"    {sl}: {desc[:80]}{'...' if len(desc) > 80 else ''}")
    print(f"  Job status:       {data['job_status']!r}")
    print(f"  Checked inclusions: {len(data['checked_inclusions'])}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
